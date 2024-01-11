from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# 1. Create new workbook
# wb = Workbook()     # To create a new workbook
wb = load_workbook("Grades2.xlsx")   # Decided to load since it already exists, hence to enable appending instead of re-writing

ws = wb.active
ws.title = 'Data'

# 2. Adding/appending data
'''
dataList1 = ["One", "Two", "Three", "Four", "Five"]
dataList2 = ["Hello","From", "Python", "To", "Excel"]
ws.append(dataList1)     # Add to cells A1, B1, C1, D1 and E1
ws.append(dataList2)     # Add to cells A2, B2, C2, D2 and E2
'''

'''
for row in range(1, 11):
    # for col in range(0, 5):
        # char = char(65 + col)
    for col in range(1, 5):
        char = get_column_letter(col)    # returns the column letters, ie 1=A, 2=B, 3=C ...
        cell_value = char+str(row)
        # print(ws[cell_value].value)
        ws[cell_value] = "Cell: " + cell_value
'''


wb.save("Grades2.xlsx")
