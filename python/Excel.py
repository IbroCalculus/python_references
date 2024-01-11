from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# NB: You cannot modify an existing workbook in python while it is open in excel
# NB: Always save at the end of the script using ie, wb.save("Grades.xlsx") to update all modifications made to the excel workbook

# 1. Load an existing workbook
wb = load_workbook("Grades.xlsx")

# 2. Accessing the active sheet from within a workbook
ws = wb.active
print(ws)       # Name of the active sheet. ie <Worksheet "Grades">

# 3. Change active worksheet title
ws.title = "Tab No. 1"

# 4. Accessing individual cell values
print(ws['A1'].value)

# 5. Modify an existing cell value in excel and saving
ws['A1'] = "Test"
print(ws['A1'].value)
wb.save("Grades.xlsx")

# 6. Get list of all sheets in the workbook
sheetList = wb.sheetnames
print(sheetList)

# 7. Accessing/working in a specific sheet
ws = wb['Sheet1']
print(ws)

# 8. Create new sheet
wb.create_sheet("TestSheet")
sheetList = wb.sheetnames
print(sheetList)
ws = wb['TestSheet']
print(ws)

# 9. Merge and unmerge cells
ws.merge_cells("A1:D4")
ws.unmerge_cells("A1:D4")

# 10. Insert empty rows and columns
ws.insert_rows(2)   # Insert new empty row at position 2
ws.insert_rows(2)   # Insert another new empty row at position 2

ws.insert_cols(2)

# 11. Delete rows and columns
ws.delete_rows(2)   # Delete row at position 2
ws.delete_cols(2)

# 12. Copying and moving cells
ws.move_range("B1:C10", rows=2, cols=5)
# row = -1 implies move UP 1 row.  row = 2 implies move DOWN 2 rows.
# col = -5 implies move LEFT 5 columns. col = 3 implies move RIGHT 3 columns

